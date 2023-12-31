import os
import openpyxl
from busqueda.buscar_empleados import (
    buscar_nombres_empleados_renacimiento,
    buscar_cedula_empleados,
    buscar_nombres_empleados_lincoln,
)


def crear_volante_empleados_renacimiento(nombre_archivo, nombre_hoja):
    wb = openpyxl.load_workbook(nombre_archivo)
    nombres = buscar_nombres_empleados_renacimiento(nombre_archivo, nombre_hoja)
    cedulas = buscar_cedula_empleados(nombre_archivo, nombre_hoja)
    if not os.path.exists("excels_renacimiento"):
        os.mkdir("excels_renacimiento")
    for dato1, dato2 in zip(cedulas, nombres):
        ws = wb["Volante Renacimiento"]
        ws["A6"] = dato1
        nombre_archivo_excel = f"excels_renacimiento/{dato2}_Volante.xlsx"
        for hoja in wb.sheetnames:
            if hoja != "Volante Renacimiento":
                wb[hoja].sheet_state = "veryHidden"
        wb.save(nombre_archivo_excel)
    wb.close()
    return None


def crear_volante_empleados_lincoln(nombre_archivo, nombre_hoja):
    wb = openpyxl.load_workbook(nombre_archivo)
    nombres = buscar_nombres_empleados_lincoln(nombre_archivo, nombre_hoja)
    cedulas = buscar_cedula_empleados(nombre_archivo, nombre_hoja)
    if not os.path.exists("excels_lincoln"):
        os.mkdir("excels_lincoln")
    for dato1, dato2 in zip(cedulas, nombres):
        ws = wb["Volante Lincoln"]
        ws["A6"] = dato1
        nombre_archivo_excel = f"excels_lincoln/{dato2}_Volante.xlsx"
        for hoja in wb.sheetnames:
            if hoja != "Volante Lincoln":
                wb[hoja].sheet_state = "veryHidden"
        wb.save(nombre_archivo_excel)
    wb.close()
    return None
