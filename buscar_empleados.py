import openpyxl


def buscar_cedula_empleados(nombre_libro, nombre_hoja):
    wb = openpyxl.load_workbook(nombre_libro, data_only=True)
    ws = wb[nombre_hoja]
    cedulas_empleados = []
    for row in ws.iter_rows(values_only=True):
        cedula_empleados = row[2]
        if cedula_empleados is not None:
            cedulas_empleados.append(cedula_empleados)
    wb.close()
    return cedulas_empleados[2:]


def buscar_nombres_empleados_renacimiento(nombre_libro, nombre_hoja):
    wb = openpyxl.load_workbook(nombre_libro, data_only=True)
    ws = wb[nombre_hoja]
    cedulas_empleados = []
    for row in ws.iter_rows(values_only=True):
        cedula_empleados = row[1]
        if cedula_empleados is not None:
            cedulas_empleados.append(cedula_empleados)
    wb.close()
    return cedulas_empleados[1:]


def buscar_nombres_empleados_lincoln(nombre_libro, nombre_hoja):
    wb = openpyxl.load_workbook(nombre_libro, data_only=True)
    ws = wb[nombre_hoja]
    cedulas_empleados = []
    for row in ws.iter_rows(values_only=True):
        cedula_empleados = row[1]
        if cedula_empleados is not None:
            cedulas_empleados.append(cedula_empleados)
    wb.close()
    return cedulas_empleados[3:]


