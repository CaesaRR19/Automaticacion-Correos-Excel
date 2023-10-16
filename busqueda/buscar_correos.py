import openpyxl


def buscar_correos_empleados_lincoln(nombre_libro, hoja_libro):
    wb = openpyxl.load_workbook(nombre_libro, data_only=True)
    ws = wb[hoja_libro]
    correos_empleados = []
    for row in ws.iter_rows(values_only=True):
        correo_empleado = row[22]
        if correo_empleado is not None:
            correos_empleados.append(correo_empleado)
    wb.close()
    return correos_empleados[1:]


def buscar_correos_empleados_renacimiento(nombre_libro, hoja_libro):
    wb = openpyxl.load_workbook(nombre_libro, data_only=True)
    ws = wb[hoja_libro]
    correos_empleados = []
    for row in ws.iter_rows(values_only=True):
        correo_empleado = row[24]
        if correo_empleado is not None:
            correos_empleados.append(correo_empleado)
    wb.close()
    return correos_empleados[1:]

